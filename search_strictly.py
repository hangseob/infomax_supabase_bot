import os
import openpyxl
import sys

def search_imdg_imdi_strictly():
    root = 'infomax_functions_templetes'
    results = []
    
    files = []
    for r, d, fs in os.walk(root):
        for f in fs:
            if f.endswith(('.xlsx', '.xlsm')) and not f.startswith('~$'):
                files.append(os.path.join(r, f))

    print(f"총 {len(files)}개 파일에서 'IMDG' 또는 'IMDI' 문자열 탐색 중...")
    
    for idx, fpath in enumerate(files, 1):
        try:
            wb = openpyxl.load_workbook(fpath, data_only=False, read_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                try:
                    for row in ws.iter_rows():
                        for cell in row:
                            val = cell.value
                            if isinstance(val, str) and ('IMDG' in val.upper() or 'IMDI' in val.upper()):
                                results.append(f"File: {fpath}\nSheet: {sn}\nCell: {cell.coordinate}\nFormula: {val}\n" + "-"*30)
                except: continue
        except: continue

    output_path = 'imdg_imdi_found.txt'
    with open(output_path, 'w', encoding='utf-8') as f:
        if results:
            f.write("\n".join(results))
            print(f"찾았습니다! {len(results)}개의 셀을 {output_path}에 저장했습니다.")
        else:
            f.write("IMDG 혹은 IMDI 문자열을 포함한 셀을 찾지 못했습니다.")
            print("발견된 셀이 없습니다.")

if __name__ == "__main__":
    search_imdg_imdi_strictly()
