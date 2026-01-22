import os
import openpyxl

def inspect_3302():
    root = 'infomax_functions_templetes'
    mkt_dir = next(d for d in os.listdir(root) if '시장분석' in d)
    stock_dir = next(d for d in os.listdir(os.path.join(root, mkt_dir)) if '주식' in d)
    file_name = next(f for f in os.listdir(os.path.join(root, mkt_dir, stock_dir)) if '3302' in f)
    path = os.path.join(root, mkt_dir, stock_dir, file_name)
    
    print(f"파일 분석 중: {path}")
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"--- 시트: {sn} ---")
        for row in ws.iter_rows(max_row=20, max_col=10):
            row_vals = [cell.value for cell in row]
            if any(isinstance(v, str) and '=' in v for v in row_vals):
                print(row_vals)

if __name__ == "__main__":
    inspect_3302()
