import os
import openpyxl

def find_all_unique_functions_in_3111():
    root = 'infomax_functions_templetes'
    mkt_dir = next(d for d in os.listdir(root) if '시장분석' in d)
    stock_dir = next(d for d in os.listdir(os.path.join(root, mkt_dir)) if '주식' in d)
    file_name = next(f for f in os.listdir(os.path.join(root, mkt_dir, stock_dir)) if '3111' in f)
    path = os.path.join(root, mkt_dir, stock_dir, file_name)
    
    print(f"파일 분석 중: {path}")
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    
    infomax_funcs = set()
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    import re
                    # Look for functions starting with IMD
                    matches = re.findall(r'IMD[A-Z0-9_]+', cell.value.upper())
                    for m in matches:
                        infomax_funcs.add(m)
    
    print(f"발견된 인포맥스 함수: {infomax_funcs}")

if __name__ == "__main__":
    find_all_unique_functions_in_3111()
