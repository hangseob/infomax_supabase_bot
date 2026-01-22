import openpyxl
import os

def find_file(root, target_name):
    for r, d, files in os.walk(root):
        for f in files:
            if target_name in f:
                return os.path.join(r, f)
    return None

file_path = find_file('infomax_functions_templetes', '6511__세계주요지수(히스토리)')
if file_path:
    print(f"Found file: {file_path}")
    wb = openpyxl.load_workbook(file_path, read_only=True)
    # Check sheet names
    print(f"Sheets: {wb.sheetnames}")
    
    # Try to find '6511(히스토리)' sheet or similar
    target_sheet = None
    for s in wb.sheetnames:
        if '6511' in s or '히스토리' in s:
            target_sheet = s
            break
    
    if target_sheet:
        ws = wb[target_sheet]
        # Row 3 for fields
        row3 = [str(cell.value) for cell in ws[3]]
        print(f"Fields (Row 3): {row3}")
        
        # Look for Vietnam (베트남) in the sheet
        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if any('베트남' in str(cell) for cell in row if cell):
                print(f"Vietnam Row ({idx}): {row}")
                break
    else:
        print("Target sheet not found.")
else:
    print("File not found.")
