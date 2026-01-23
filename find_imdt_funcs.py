import os
import openpyxl
from pyxlsb import open_workbook as open_xlsb
import traceback

def search_imdt_in_excel(root_dir):
    results = []
    
    for root, dirs, files in os.walk(root_dir):
        for file in files:
            if file.startswith('~$'): continue
            
            file_path = os.path.join(root, file)
            ext = os.path.splitext(file).lower()
            
            if ext in ['.xlsx', '.xlsm']:
                try:
                    # data_only=False to read formulas
                    wb = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
                    for sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        # Use try-except for sheet iteration as some sheets might be problematic
                        try:
                            for row in sheet.iter_rows():
                                for cell in row:
                                    if cell.value and isinstance(cell.value, str) and 'IMDT' in cell.value.upper():
                                        results.append({
                                            'file': file_path,
                                            'sheet': sheet_name,
                                            'cell': cell.coordinate,
                                            'formula': cell.value
                                        })
                                        # One match per sheet is enough for this request, or keep going? 
                                        # User asked for "Excel name and sheet name", so one per sheet is enough.
                                        break
                                else: continue
                                break
                        except Exception as e:
                            print(f"Error reading sheet {sheet_name} in {file}: {e}")
                    wb.close()
                except Exception as e:
                    print(f"Error reading {file}: {e}")
                    
            elif ext == '.xlsb':
                try:
                    with open_xlsb(file_path) as wb:
                        for sheet_name in wb.sheets:
                            with wb.get_sheet(sheet_name) as sheet:
                                for row in sheet.rows():
                                    for cell in row:
                                        # pyxlsb formula handling is different
                                        # but often formulas are stored as strings starting with '='
                                        if cell.v and isinstance(cell.v, str) and 'IMDT' in cell.v.upper():
                                            results.append({
                                                'file': file_path,
                                                'sheet': sheet_name,
                                                'cell': f"R{cell.r}C{cell.c}",
                                                'formula': cell.v
                                            })
                                            break
                                    else: continue
                                    break
                except Exception as e:
                    print(f"Error reading {file}: {e}")
                    
    return results

if __name__ == "__main__":
    target_dir = r"infomax_functions_templetes"
    print(f"Searching for 'IMDT' in {target_dir}...")
    
    matches = search_imdt_in_excel(target_dir)
    
    if matches:
        print(f"\nFound {len(matches)} occurrences of IMDT:")
        current_file = ""
        for match in matches:
            if match['file'] != current_file:
                print(f"\n[File: {match['file']}]")
                current_file = match['file']
            print(f"  - Sheet: {match['sheet']} (Cell: {match['cell']}, Formula: {match['formula']})")
    else:
        print("\nNo IMDT functions found.")
