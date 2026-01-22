import os
import openpyxl

def inspect_3206_filter_columns():
    root = 'infomax_functions_templetes'
    mkt_dir = next(d for d in os.listdir(root) if '시장분석' in d)
    stock_dir = next(d for d in os.listdir(os.path.join(root, mkt_dir)) if '주식' in d)
    file_name = next(f for f in os.listdir(os.path.join(root, mkt_dir, stock_dir)) if '3206' in f)
    path = os.path.join(root, mkt_dir, stock_dir, file_name)
    
    print(f"파일 분석 중: {path}")
    wb = openpyxl.load_workbook(path, data_only=False)
    
    if 'filter' in wb.sheetnames:
        ws = wb['filter']
        print(f"--- '{ws.title}' 시트 I열, J열 확인 (I1:J20) ---")
        for row in ws.iter_rows(min_row=1, max_row=20, min_col=9, max_col=10):
            print(f"  {[cell.coordinate for cell in row]}: {[cell.value.text if isinstance(cell.value, openpyxl.worksheet.formula.ArrayFormula) else cell.value for cell in row]}")

if __name__ == "__main__":
    inspect_3206_filter_columns()
