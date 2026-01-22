import os
import openpyxl
import sys

# Ensure UTF-8 output
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def inspect_imdt_usage():
    root = 'infomax_functions_templetes'
    files_with_imdt = [
        '기업분석/4212__신용등급_변동.xlsm',
        '시장분석/채권/4336_4345__CD_CP_Term.xlsx',
        '시장분석/채권/4541__종목별_시가평가_계산.xlsx'
    ]
    
    for rel_path in files_with_imdt:
        parts = rel_path.split('/')
        curr = root
        for p in parts:
            try:
                matches = [d for d in os.listdir(curr) if p in d or d in p]
                if matches: curr = os.path.join(curr, matches[0])
                else: curr = None; break
            except: curr = None; break
            
        if curr and os.path.exists(curr):
            print(f"\n파일 분석: {rel_path}")
            wb = openpyxl.load_workbook(curr, data_only=False, read_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                try:
                    for row in ws.iter_rows():
                        for cell in row:
                            if isinstance(cell.value, str) and 'IMDT' in cell.value.upper():
                                print(f"  [{sn} {cell.coordinate}] {cell.value}")
                except: continue

if __name__ == "__main__":
    inspect_imdt_usage()
