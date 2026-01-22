import os
import openpyxl
import re

def extract_function_names():
    root = 'infomax_functions_templetes'
    functions = set()
    
    # 템플릿 파일 몇 개를 샘플링해서 함수명을 추출합니다.
    sample_files = [
        '시장분석/주식/3111__주식종목정보.xlsx',
        '시장분석/주식/3206__KRX_지수_종합.xlsx',
        '경제분석/국내_거시경제지표.xlsx'
    ]
    
    print("인포맥스 함수명 추출 중...")
    
    for rel_path in sample_files:
        # 실제 경로 찾기 (한글 인코딩 문제 대응)
        found_path = None
        target_parts = rel_path.split('/')
        curr = root
        for part in target_parts:
            try:
                matches = [d for d in os.listdir(curr) if part in d or d in part]
                if matches:
                    curr = os.path.join(curr, matches[0])
                else:
                    curr = None
                    break
            except:
                curr = None
                break
        
        if curr and os.path.exists(curr):
            print(f"파일 분석: {rel_path}")
            try:
                wb = openpyxl.load_workbook(curr, data_only=False, read_only=True)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(max_row=100, max_col=20):
                        for cell in row:
                            val = cell.value
                            if isinstance(val, str) and val.startswith('='):
                                # 함수명 추출 (예: =IMDP(...) -> IMDP)
                                matches = re.findall(r'([A-Z][A-Z0-9_]+)\(', val.upper())
                                for m in matches:
                                    functions.add(m)
            except Exception as e:
                print(f"  오류: {e}")
    
    print("\n[발견된 함수 목록]")
    for func in sorted(list(functions)):
        print(f"- {func}")

if __name__ == "__main__":
    extract_function_names()
