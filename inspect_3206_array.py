import os
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

def inspect_3206_array_formulas():
    root = 'infomax_functions_templetes'
    mkt_dir = next(d for d in os.listdir(root) if '시장분석' in d)
    stock_dir = next(d for d in os.listdir(os.path.join(root, mkt_dir)) if '주식' in d)
    file_name = next(f for f in os.listdir(os.path.join(root, mkt_dir, stock_dir)) if '3206' in f)
    path = os.path.join(root, mkt_dir, stock_dir, file_name)
    
    print(f"파일 분석 중: {path}")
    wb = openpyxl.load_workbook(path, data_only=False)
    
    if 'KRX' in wb.sheetnames:
        ws = wb['KRX']
        print(f"--- '{ws.title}' 시트의 배열 수식 확인 (D9:E15) ---")
        for row in ws.iter_rows(min_row=9, max_row=15, min_col=4, max_col=5):
            for cell in row:
                if isinstance(cell.value, ArrayFormula):
                    print(f"  {cell.coordinate}: [Array] {cell.value.text}")
                else:
                    print(f"  {cell.coordinate}: {cell.value}")

if __name__ == "__main__":
    inspect_3206_array_formulas()
