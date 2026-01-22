import openpyxl
import os

def search_market_cap_field(root_dir):
    search_terms = ['시가총액', '시총', 'MarketCap', 'Mkt Cap']
    results = []
    
    for r, d, files in os.walk(root_dir):
        for file in files:
            if file.endswith(('.xlsx', '.xlsm')):
                file_path = os.path.join(r, file)
                try:
                    # Search in formulas and values
                    wb = openpyxl.load_workbook(file_path, read_only=True)
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        for row in ws.iter_rows(values_only=False):
                            for cell in row:
                                if cell.value and isinstance(cell.value, str):
                                    val = cell.value
                                    if any(term.lower() in val.lower() for term in search_terms):
                                        results.append({
                                            'file': file,
                                            'sheet': sheet_name,
                                            'cell': cell.coordinate,
                                            'value': val
                                        })
                                        # Stop after finding one example in a file to be efficient
                                        break
                            else: continue
                            break
                except:
                    pass
    return results

if __name__ == "__main__":
    found = search_market_cap_field('infomax_functions_templetes')
    for item in found:
        print(f"File: {item['file']} | Sheet: {item['sheet']} | Cell: {item['cell']}")
        print(f"  Value: {item['value']}")
        print("-" * 30)
