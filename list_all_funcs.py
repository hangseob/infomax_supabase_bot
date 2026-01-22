import os
import openpyxl
import re

def list_all_imd_functions():
    root = 'infomax_functions_templetes'
    all_imd_funcs = set()
    
    files_to_check = []
    for r, d, filenames in os.walk(root):
        for f in filenames:
            if f.endswith(('.xlsx', '.xlsm')) and not f.startswith('~$'):
                files_to_check.append(os.path.join(r, f))

    print(f"총 {len(files_to_check)}개의 파일에서 인포맥스 함수명 추출 중...")
    
    for idx, file_path in enumerate(files_to_check, 1):
        try:
            wb = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                try:
                    for row in ws.iter_rows():
                        for cell in row:
                            if isinstance(cell.value, str) and '=' in cell.value:
                                matches = re.findall(r'IMD[A-Z0-9_]+', cell.value.upper())
                                for m in matches:
                                    all_imd_funcs.add(m)
                except: continue
        except: continue
        
        if idx % 20 == 0:
            print(f"  {idx}/{len(files_to_check)} 파일 완료...")

    print("\n[발견된 모든 인포맥스 함수 목록]")
    for func in sorted(list(all_imd_funcs)):
        print(f"- {func}")

if __name__ == "__main__":
    list_all_imd_functions()
