import os
import openpyxl

def inspect_6535():
    root = 'infomax_functions_templetes'
    mkt_dir = next(d for d in os.listdir(root) if '시장분석' in d)
    overseas_dir = next(d for d in os.listdir(os.path.join(root, mkt_dir)) if '해외' in d)
    file_name = next(f for f in os.listdir(os.path.join(root, mkt_dir, overseas_dir)) if '6535' in f)
    path = os.path.join(root, mkt_dir, overseas_dir, file_name)
    
    print(f"파일 분석 중: {path}")
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"--- 시트: {sn} ---")
        for row in ws.iter_rows(max_row=50, max_col=15):
            for cell in row:
                if isinstance(cell.value, str) and 'IMD' in cell.value.upper():
                    print(f"  [{cell.coordinate}] {cell.value}")

if __name__ == "__main__":
    inspect_6535()
