import os
import openpyxl
import re
import sys

def resolve_with_openpyxl(formula, value_sheet):
    if not isinstance(formula, str) or not formula.startswith('='):
        return formula
    cell_pattern = r"(\$?[A-Z]+\$?\d+)"
    def replace_match(match):
        cell_ref = match.group(1).replace('$', '')
        try:
            val = value_sheet[cell_ref].value
            if val is None: return "None"
            if isinstance(val, str): return f'"{val}"'
            return str(val)
        except: return match.group(1)
    return re.sub(cell_pattern, replace_match, formula)

def find_all_imd_functions():
    root = 'infomax_functions_templetes'
    output_path = 'imd_all_results.txt'
    
    print("\n[진행 상황] 모든 인포맥스 수식(IMD로 시작)을 탐색합니다...")
    
    files_to_check = []
    for r, d, filenames in os.walk(root):
        for f in filenames:
            if f.endswith(('.xlsx', '.xlsm')) and not f.startswith('~$'):
                files_to_check.append(os.path.join(r, f))

    total_files = len(files_to_check)
    results = []
    
    for idx, file_path in enumerate(files_to_check, 1):
        filename = os.path.basename(file_path)
        print(f"[{idx}/{total_files}] 검사 중: {filename}...", end='\r')
        sys.stdout.flush()
        
        try:
            wb_f = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
            wb_v = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            
            file_found_count = 0
            for sn in wb_f.sheetnames:
                ws_f = wb_f[sn]
                ws_v = wb_v[sn]
                try:
                    for row in ws_f.iter_rows():
                        for cell in row:
                            val = cell.value
                            if isinstance(val, str) and 'IMD' in val.upper():
                                resolved = resolve_with_openpyxl(val, ws_v)
                                res = f"File: {file_path}\nSheet: {sn}\nCell: {cell.coordinate}\nOriginal: {val}\nResolved: {resolved}\n" + "-"*50 + "\n"
                                results.append(res)
                                file_found_count += 1
                except: continue
            
            if file_found_count > 0:
                print(f"\n[*] {filename}: {file_found_count}개 발견 (누적 {len(results)}개)")
                
        except Exception:
            continue

    with open(output_path, 'w', encoding='utf-8') as f_out:
        if results:
            f_out.writelines(results)
        else:
            f_out.write("인포맥스 수식을 찾지 못했습니다.")
            
    print(f"\n\n[탐색 완료] 총 {len(results)}개의 수식을 {output_path}에 저장했습니다.")

if __name__ == "__main__":
    find_all_imd_functions()
