import os
import openpyxl

def inspect_education_materials():
    root = 'infomax_functions_templetes'
    edu_dir = next(d for d in os.listdir(root) if '교육자료' in d)
    
    print(f"교육자료 폴더 분석 중: {edu_dir}")
    for f in os.listdir(os.path.join(root, edu_dir)):
        if f.endswith(('.xlsx', '.xlsm')) and not f.startswith('~$'):
            path = os.path.join(root, edu_dir, f)
            print(f"\n--- 파일: {f} ---")
            try:
                wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
                for sn in wb.sheetnames:
                    ws = wb[sn]
                    # 샘플로 몇 개만 확인
                    for row in ws.iter_rows(max_row=50, max_col=10):
                        for cell in row:
                            if isinstance(cell.value, str) and 'IMD' in cell.value.upper():
                                print(f"  [{sn} {cell.coordinate}] {cell.value}")
            except: continue

if __name__ == "__main__":
    inspect_education_materials()
