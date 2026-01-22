import os
import openpyxl

def inspect_3302_details():
    root = 'infomax_functions_templetes'
    mkt_dir = next(d for d in os.listdir(root) if '시장분석' in d)
    stock_dir = next(d for d in os.listdir(os.path.join(root, mkt_dir)) if '주식' in d)
    file_name = next(f for f in os.listdir(os.path.join(root, mkt_dir, stock_dir)) if '3302' in f)
    path = os.path.join(root, mkt_dir, stock_dir, file_name)
    
    wb = openpyxl.load_workbook(path, data_only=True) # Read values
    ws = wb['Market Overview']
    print(f"Market Overview N2 value: {ws['N2'].value}")
    
    print("\n--- Sheets in 3302 ---")
    print(wb.sheetnames)
    
    for sn in wb.sheetnames:
        if 'Code' in sn or '코드' in sn:
            ws_code = wb[sn]
            print(f"\n--- {sn} sheet (A1:B20) ---")
            for row in ws_code.iter_rows(max_row=20, max_col=2):
                print([cell.value for cell in row])

if __name__ == "__main__":
    inspect_3302_details()
