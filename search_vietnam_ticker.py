import openpyxl
import os

def search_excel(file_path, search_terms):
    try:
        # data_only=True to get calculated values if any, though formulas are more likely
        wb = openpyxl.load_workbook(file_path, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Search both values and formulas
            for row_idx, row in enumerate(ws.iter_rows(values_only=False), 1):
                for col_idx, cell in enumerate(row, 1):
                    cell_val = str(cell.value) if cell.value else ""
                    for term in search_terms:
                        if term in cell_val:
                            # If found, let's also look at nearby cells to find the ticker
                            nearby = []
                            # Look at current row
                            for r_cell in row:
                                nearby.append(str(r_cell.value))
                            return True, sheet_name, cell.coordinate, cell_val, nearby
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
    return False, None, None, None, None

files_to_check = [
    'infomax_functions_templetes/시장분석/해외/6511__세계주요지수(미니).xlsx',
    'infomax_functions_templetes/시장분석/해외/6511__세계주요지수(히스토리).xlsx',
    'infomax_functions_templetes/시장분석/해외/6535__국가별_금융종합.xlsm',
    'infomax_functions_templetes/경제분석/세계주요증시_및_상품시세.xlsx',
    'infomax_functions_templetes/경제분석/국내시장_주요지표.xlsx'
]

search_terms = ['베트남', 'Vietnam', '호치민', 'Ho Chi Minh', 'VNI', 'VNC']

print("Searching for Vietnam index information...")
for f_path in files_to_check:
    if not os.path.exists(f_path):
        continue
    print(f"Checking {f_path}...")
    found, sheet, coord, val, nearby = search_excel(f_path, search_terms)
    if found:
        print(f"\n[FOUND MATCH]")
        print(f"File: {f_path}")
        print(f"Sheet: {sheet}")
        print(f"Coordinate: {coord}")
        print(f"Value: {val}")
        print(f"Nearby values in row: {nearby}")
        print("-" * 30)

print("\nSearch complete.")
