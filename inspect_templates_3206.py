import os
import openpyxl

def inspect_3206_krx():
    root = 'infomax_functions_templetes'
    # Find 3206 file
    mkt_dir = next(d for d in os.listdir(root) if '시장분석' in d)
    stock_dir = next(d for d in os.listdir(os.path.join(root, mkt_dir)) if '주식' in d)
    file_name = next(f for f in os.listdir(os.path.join(root, mkt_dir, stock_dir)) if '3206' in f)
    path = os.path.join(root, mkt_dir, stock_dir, file_name)
    
    print(f"파일 분석 중: {path}")
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    
    # KRX 시트 확인
    if 'KRX' in wb.sheetnames:
        ws = wb['KRX']
        print(f"--- '{ws.title}' 시트 내용 (A8:E30) ---")
        for row in ws.iter_rows(min_row=8, max_row=30, min_col=1, max_col=5):
            print([cell.value for cell in row])
            
    # filter 시트 확인
    if 'filter' in wb.sheetnames:
        ws = wb['filter']
        print(f"\n--- '{ws.title}' 시트 내용 (A1:E20) ---")
        for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=5):
            print([cell.value for cell in row])

if __name__ == "__main__":
    inspect_3206_krx()
