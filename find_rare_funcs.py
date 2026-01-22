import os
import openpyxl
import sys

def find_rare_imd_functions():
    root = 'infomax_functions_templetes'
    rare_funcs = {}
    
    files_to_check = []
    for r, d, filenames in os.walk(root):
        for f in filenames:
            if f.endswith(('.xlsx', '.xlsm')) and not f.startswith('~$'):
                files_to_check.append(os.path.join(r, f))

    print(f"총 {len(files_to_check)}개의 파일에서 희귀 인포맥스 함수 탐색 중...")
    
    for idx, file_path in enumerate(files_to_check, 1):
        try:
            wb = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                try:
                    for row in ws.iter_rows():
                        for cell in row:
                            if isinstance(cell.value, str) and cell.value.startswith('='):
                                import re
                                matches = re.findall(r'IMD[A-Z0-9_]+', cell.value.upper())
                                for m in matches:
                                    if m not in ['IMDP', 'IMDH']:
                                        if m not in rare_funcs:
                                            rare_funcs[m] = []
                                        rare_funcs[m].append((file_path, sheet, cell.coordinate, cell.value))
                except: continue
        except: continue
        
        if idx % 10 == 0:
            print(f"  {idx}/{len(files_to_check)} 파일 완료...")

    print("\n[발견된 희귀 인포맥스 함수]")
    for func, occurrences in rare_funcs.items():
        print(f"\n함수: {func} ({len(occurrences)}회 발견)")
        for f, s, c, val in occurrences[:3]: # 상위 3개만 출력
            print(f"  파일: {f}\n  시트: {s}\n  셀: {c}\n  수식: {val}")

if __name__ == "__main__":
    find_rare_imd_functions()
