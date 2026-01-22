import os

def find_file(root, filename):
    for r, d, files in os.walk(root):
        for f in files:
            if filename in f:
                return os.path.join(r, f)
    return None

target_file = "우리집 가계 금융 현황.종합.xlsx"
onedrive_path = r"C:\Users\hangs\OneDrive"

print(f"Searching for '{target_file}' in '{onedrive_path}'...")
result = find_file(onedrive_path, target_file)

if result:
    print(f"Found: {result}")
    # Now try to read it with openpyxl
    import openpyxl
    try:
        # We only need to check table columns, so we can try to find the sheet with the table
        wb = openpyxl.load_workbook(result, read_only=True, data_only=True)
        print(f"Loaded workbook. Sheets: {wb.sheetnames}")
        
        # Unfortunately openpyxl's read_only mode doesn't support table definitions well.
        # But we can try to find the headers manually if we know the table name.
        # Actually, let's try to find where "표.거래내역" might be.
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Tables are stored in ws._tables in some versions, but let's just search for common headers
            # or the table name in the sheet's metadata if possible.
            # Since we can't easily get table headers from openpyxl without loading everything,
            # let's just look at the first few rows of each sheet.
            print(f"Checking sheet: {sheet_name}")
            for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                if row and any(row):
                    print(f"  Row: {row}")
    except Exception as e:
        print(f"Error reading file: {e}")
else:
    print("File not found.")
