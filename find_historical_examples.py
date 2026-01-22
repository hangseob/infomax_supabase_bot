import openpyxl
import os
import re

def search_historical_examples(root_dir):
    results = []
    # Relevant patterns for historical data and overseas indices
    # IMDH is the main historical function
    target_funcs = ['IMDH', 'IMDT'] 
    
    print(f"Searching for historical data function examples in {root_dir}...")
    
    for r, d, files in os.walk(root_dir):
        for file in files:
            if file.endswith(('.xlsx', '.xlsm')):
                file_path = os.path.join(r, file)
                try:
                    wb = openpyxl.load_workbook(file_path, read_only=True)
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        # We only need a few examples per file
                        examples_found = 0
                        for row in ws.iter_rows(values_only=False):
                            for cell in row:
                                if cell.value and isinstance(cell.value, str) and '=' in cell.value:
                                    formula = cell.value.upper()
                                    if any(f in formula for f in target_funcs):
                                        # Heuristic: Check if it's likely an overseas index
                                        # Overseas indices often use "IDX", "FRN", or specific ticker formats
                                        if any(x in formula for x in ['"IDX"', '"FRN"', '"STK"', 'VNI', 'HSI', 'N225', 'DJI', 'SPX']):
                                            results.append({
                                                'file': file,
                                                'sheet': sheet_name,
                                                'cell': cell.coordinate,
                                                'formula': cell.value
                                            })
                                            examples_found += 1
                                            if examples_found >= 3: # Limit per file
                                                break
                            if examples_found >= 3:
                                break
                except Exception as e:
                    # print(f"Error reading {file}: {e}")
                    pass
    return results

if __name__ == "__main__":
    examples = search_historical_examples('infomax_functions_templetes')
    print(f"\nFound {len(examples)} examples:\n")
    for ex in examples:
        print(f"File: {ex['file']}")
        print(f"Sheet: {ex['sheet']} | Cell: {ex['cell']}")
        print(f"Formula: {ex['formula']}")
        print("-" * 50)
