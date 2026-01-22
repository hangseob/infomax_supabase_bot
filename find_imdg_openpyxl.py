import os
import openpyxl
import re
import sys

def resolve_with_openpyxl(formula, value_sheet):
    """
    수식 내의 셀 참조를 value_sheet에서 찾은 실제 값으로 치환합니다.
    """
    if not isinstance(formula, str) or not formula.startswith('='):
        return formula

    # 간단한 셀 참조 패턴 ($A$1, A1 등)
    cell_pattern = r"(\$?[A-Z]+\$?\d+)"
    
    def replace_match(match):
        cell_ref = match.group(1).replace('$', '')
        try:
            val = value_sheet[cell_ref].value
            if val is None: return "None"
            if isinstance(val, str): return f'"{val}"'
            return str(val)
        except:
            return match.group(1)

    return re.sub(cell_pattern, replace_match, formula)

def find_imdg_openpyxl():
    root = 'infomax_functions_templetes'
    output_path = 'imdg_imdi_results.txt'
    
    print("\n[진행 상황] openpyxl을 사용하여 수식을 탐색합니다 (엑셀 실행 불필요)...")
    
    files_to_check = []
    for r, d, filenames in os.walk(root):
        for f in filenames:
            # openpyxl은 xlsx, xlsm 지원 (xlsb는 지원 안 함)
            if f.endswith(('.xlsx', '.xlsm')) and not f.startswith('~$'):
                files_to_check.append(os.path.join(r, f))

    total_files = len(files_to_check)
    print(f"-> 총 {total_files}개의 파일을 검사합니다.\n")

    results = []
    
    for idx, file_path in enumerate(files_to_check, 1):
        filename = os.path.basename(file_path)
        print(f"[{idx}/{total_files}] {filename} 읽는 중...", end='\r')
        sys.stdout.flush()
        
        try:
            # 1. 수식 포함 로드
            wb_formula = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
            # 2. 값 포함 로드
            wb_value = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            
            found_in_file = 0
            for sheet_name in wb_formula.sheetnames:
                ws_f = wb_formula[sheet_name]
                ws_v = wb_value[sheet_name]
                
                # 모든 셀 순회
                for row in ws_f.iter_rows():
                    for cell in row:
                        formula = cell.value
                        if isinstance(formula, str) and formula.startswith('='):
                            f_upper = formula.upper()
                            if "IMDG" in f_upper or "IMDI" in f_upper:
                                # 셀 참조 해결
                                resolved = resolve_with_openpyxl(formula, ws_v)
                                
                                res = f"File: {file_path}\n"
                                res += f"Sheet: {sheet_name}\n"
                                res += f"Cell: {cell.coordinate}\n"
                                res += f"Original: {formula}\n"
                                res += f"Resolved: {resolved}\n"
                                res += "-" * 50 + "\n"
                                results.append(res)
                                found_in_file += 1
            
            if found_in_file > 0:
                print(f"\n[*] {filename}: {found_in_file}개 발견! (누적 {len(results)}개)")
                
        except Exception as e:
            # xlsm 등 특정 파일에서 에러가 날 수 있음
            print(f"\n[!] {filename} 오류: {e}")
            continue

    # 결과 저장
    with open(output_path, 'w', encoding='utf-8') as f_out:
        if results:
            f_out.writelines(results)
        else:
            f_out.write("검사 결과 IMDG 또는 IMDI 수식을 사용하는 셀을 찾지 못했습니다.")
            
    print(f"\n\n[탐색 완료]")
    print(f"- 검사한 파일: {total_files}개")
    print(f"- 발견한 수식: {len(results)}개")
    print(f"- 결과 파일: {output_path}")

if __name__ == "__main__":
    find_imdg_openpyxl()
